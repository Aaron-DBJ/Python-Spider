# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.webdriver.support.wait import WebDriverWait
import time
from bs4 import BeautifulSoup
import json
import requests
from threading import Thread
import pymysql
import openpyxl

begin = time.time()
header = {
    "User-Agent" : "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_5) "
        + "AppleWebKit/537.36 (KHTML, like Gecko) "
        + "Chrome/67.0.3396.99 Safari/537.36"
}
images = []
names = []



def get_page(url):
    repsonse = requests.get(url, headers=header)
    if repsonse.status_code == 200:
        return repsonse.text
    return None


def parse_page(html):
    soup = BeautifulSoup(html, 'lxml')
    novels = soup.find('div', class_="book-img-text").find('ul').find_all('li')
    for item in novels:
        yield{
            'rank': item.find('span', class_="rank-tag").get_text(),
            'title': item.find('div', class_="book-mid-info").find('a').get_text().strip(),
            'author': item.find('p', class_='author').find('a', class_="name").get_text().strip(),
            'intro': item.find('p', class_="intro").get_text().strip()
        }
        image = "https:" + item.find('div', class_="book-img-box").find('img').get('src')
        name = item.find('div', class_="book-mid-info").find('a').get_text()
        images.append(image)
        names.append(name)
    

def main(page):
    url = "https://www.qidian.com/rank/hotsales?chn=5&page={}".format(page)
    html = get_page(url)
    novels = parse_page(html)
    print('连接数据库...')
    # 连接数据库
    try:
        conn = pymysql.connect(
            host='localhost', 
            user='root', 
            password='root', 
            port=3306, 
            db='spiders', 
            charset='utf8'
        )
        print('数据库连接成功!')
        # 使用cursor（）方法获取操作游标
        cursor = conn.cursor()
    except Exception as e:
        print('连接数据库失败.')
        print(e)
    for item in novels:
        save_in_excel(item)
        save_in_db(item, conn, cursor)


def save_in_db(item, conn, cursor):
    insert = "INSERT INTO Histroynovel VALUES('{}','{}','{}','{}')".format(item['rank'], item['title'], item['author'], item['intro'])
    try:
        cursor.execute(insert)
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(e)


def download_image(images):
    i = 0
    path = '/Users/aaron_dbj/PythonProjects/爬虫图片/历史小说封面/'
    for image_url in images:
        image = requests.get(image_url, headers=header).content
        file_name = path + names[i] + ".jpg"
        print("《{}》 is downloading.".format(names[i]))
        with open(file_name, 'wb') as f:
            f.write(image)
        i += 1


wb = openpyxl.Workbook()


def save_in_excel(content):
    ws = wb.active
    ws['A1'] = '排名'
    ws['B1'] = '书名'
    ws['C1'] = '作者'
    ws['D1'] = '简介'
    rank = content['rank']
    title = content['title']
    author = content['author']
    intro = content['intro']
    ws.append([rank, title, author, intro])


if __name__ == "__main__":
    for i in range(1, 11):
        main(i)
    download_image(images)
    wb.save("/Users/aaron_dbj/PythonProjects/爬虫Excel/历史小说排行.xlsx")
    end = time.time()
    print("Processing time : {}".format(end-begin))
