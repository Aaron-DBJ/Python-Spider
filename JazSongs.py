import requests
from bs4 import BeautifulSoup
import time
import openpyxl
import pymysql
from threading import Thread
import json

header = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6)'
    + 'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/'
    + '68.0.3440.75 Safari/537.36'
}
start = time.time()

def get_page(url):
    response = requests.get(url, headers=header)
    if response.status_code == 200:
        return response.text
    return None


def parse_page(html):
    soup = BeautifulSoup(html, 'lxml')
    songs = soup.find('table', class_='track_list').find_all('tr')
    for song in songs:
        yield{
            'rank': song.find('td', class_='trackid').get_text().strip(),
            'name': song.find('td', class_='song_name').get_text().strip(),
            'hotratio': song.find('td', class_='song_hot').get_text().strip()
        }


def write_to_file(content):
    with open('爬虫文本/周杰伦歌曲大全.txt', 'a+', encoding='utf-8') as f:
        f.write(json.dumps(content, ensure_ascii=False)+'\n')


def main(spm, offset):
    url = 'https://www.xiami.com/artist/top-1260?spm={}&page={}'.format(spm, offset)
    html = get_page(url)
    for item in parse_page(html):
        print(item)
        write_to_file(item)
        save_in_excel(item)
        save_in_mysql(conn, cursor, item)


wb = openpyxl.Workbook()


def save_in_excel(item):
    ws = wb.active
    ws['A1'] = '排名'
    ws['B1'] = '歌名'
    ws['C1'] = '热度'
    ws.append([item['rank'], item['name'], item['hotratio']])


def save_in_mysql(conn, cursor, item):
    sql = "INSERT INTO jaysong VALUES('{}', '{}', '{}')".format(item['rank'], item['name'], item['hotratio'])
    try:
        cursor.execute(sql)
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(e)


if __name__ == '__main__':
    print('连接数据库......')
    try:
        conn = pymysql.connect(
            host='localhost', 
            user='root', 
            password='root', 
            port=3306, 
            db='spiders', 
            charset='utf8'
        )
        cursor = conn.cursor()
        print('数据库连接成功......')
    except Exception as e:
        print('连接失败......')
        print(e)
    for i in range(1, 6):
        t = Thread(target=main, args=('0.0.0.0.xckaF1', i))
        # main('0.0.0.0.xckaF1', i)
        t.start()
        t.join()
    wb.save('爬虫Excel/周杰伦歌曲大全.xlsx')
    print('###################-Completed-###################')
    end = time.time()
    print("Processing time: {}".format(end - start))